# =====================================================================
# Informe en Excel — pensado para leerse, no para volcarse.
# Antes eran 3 hojas con los dataframes crudos: sin formato, con nombres
# técnicos y sin resumen. Quien lo abría veía datos, no conclusiones.
# Ahora abre en un resumen ejecutivo y el detalle queda detrás.
# =====================================================================
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Paleta de marca (la misma del informe en pantalla).
NAVY, ORO, TEAL, ROJO = "0C1A2E", "C8901C", "28B088", "E05252"
GRIS, GRIS_CLARO, BLANCO = "5A7391", "EEF3FA", "FFFFFF"

_TONO_COLOR = {"alerta": ORO, "ok": TEAL, "info": "2B6CB0", "accion": "2B6CB0"}
_LINEA = Side(style="thin", color="DDE6F0")


def _celda(ws, fila, col, valor, *, negrita=False, tam=11, color="1A2B45",
           fondo=None, wrap=False, formato=None, arriba=False):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = Font(name="Calibri", size=tam, bold=negrita, color=color)
    if fondo:
        c.fill = PatternFill("solid", fgColor=fondo)
    c.alignment = Alignment(wrap_text=wrap, vertical="top" if (wrap or arriba) else "center")
    if formato:
        c.number_format = formato
    return c


def _titulo_hoja(ws, texto, subtitulo=None, ancho=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    _celda(ws, 1, 1, texto, negrita=True, tam=16, color=BLANCO, fondo=NAVY)
    ws.row_dimensions[1].height = 30
    if subtitulo:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho)
        _celda(ws, 2, 1, subtitulo, tam=10, color=GRIS, wrap=True)
        ws.row_dimensions[2].height = 26
    return 4 if subtitulo else 3


def _cabecera_tabla(ws, fila, encabezados, anchos):
    for i, (h, w) in enumerate(zip(encabezados, anchos), start=1):
        _celda(ws, fila, i, h, negrita=True, tam=9, color=BLANCO, fondo=GRIS)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


def _bordear(ws, f1, f2, ncols):
    for f in range(f1, f2 + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=f, column=c).border = Border(bottom=_LINEA)


# ---------------------------------------------------------------------
def _hoja_resumen(wb, inf):
    """Lo que un gerente lee: veredicto, qué hacer y las cifras clave."""
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    k = inf.get("kpis", {})
    a = inf.get("asistente", {}) or {}
    u = k.get("unidad", "unidades")

    f = _titulo_hoja(ws, "Informe de planificación",
                     f"Proyección de {k.get('horizonte','')} periodos "
                     f"{k.get('frecuencia_plural','')} · {k.get('n_series','')} series · "
                     f"generado por Forecast de automatizaesto", ancho=6)
    for col, w in zip("ABCDEF", (34, 20, 20, 20, 20, 20)):
        ws.column_dimensions[col].width = w

    # --- Estado del negocio ---
    ver = a.get("veredicto") or {}
    if ver.get("frase"):
        _celda(ws, f, 1, "ESTADO DEL NEGOCIO", negrita=True, tam=9, color=GRIS)
        f += 1
        ws.merge_cells(start_row=f, start_column=1, end_row=f + 2, end_column=6)
        _celda(ws, f, 1, ver["frase"], negrita=True, tam=13, wrap=True,
               fondo=GRIS_CLARO, color=NAVY)
        ws.row_dimensions[f].height = 24
        f += 4

    # --- Cifras clave ---
    _celda(ws, f, 1, "CIFRAS CLAVE", negrita=True, tam=9, color=GRIS)
    f += 1
    cifras = [("Proyección total", f"{k.get('total_fmt','')} {u}"),
              ("Escenario bajo (comprometer)", f"{k.get('piso_fmt','')} {u}"),
              ("Escenario alto (preparar)", f"{k.get('techo_fmt','')} {u}"),
              ("Variación vs periodo anterior", f"{k.get('variacion_pct',0):+.1f}%"),
              ("Margen de error validado", f"{k.get('wape_pct',0)}%"),
              ("Nivel de confianza", str(k.get("confianza", "")).title())]
    for i, (lab, val) in enumerate(cifras):
        col = 1 + (i % 3) * 2
        fil = f + (i // 3) * 2
        _celda(ws, fil, col, lab, tam=9, color=GRIS)
        _celda(ws, fil + 1, col, val, negrita=True, tam=13, color=NAVY)
    f += 5

    # --- Qué deberías hacer ---
    dec = a.get("decisiones") or []
    if dec:
        _celda(ws, f, 1, "QUÉ DEBERÍAS HACER", negrita=True, tam=9, color=GRIS)
        f += 1
        _cabecera_tabla(ws, f, ["Recomendación", "Prioridad", "Qué hacer", "", "Por qué importa", ""],
                        [34, 12, 30, 18, 34, 18])
        ws.freeze_panes = None
        f += 1
        ini = f
        for d in dec:
            color = _TONO_COLOR.get(d.get("tono"), GRIS)
            _celda(ws, f, 1, d.get("titulo", ""), negrita=True, wrap=True, tam=10, color=NAVY)
            _celda(ws, f, 2, (d.get("prioridad") or "").upper(), negrita=True, tam=9, color=color)
            ws.merge_cells(start_row=f, start_column=3, end_row=f, end_column=4)
            _celda(ws, f, 3, d.get("accion", ""), wrap=True, tam=10)
            ws.merge_cells(start_row=f, start_column=5, end_row=f, end_column=6)
            _celda(ws, f, 5, d.get("importa", ""), wrap=True, tam=9, color=GRIS)
            ws.row_dimensions[f].height = 46
            f += 1
        _bordear(ws, ini, f - 1, 6)
        f += 2

    # --- Plan de acción ---
    plan = a.get("plan_accion") or []
    if plan:
        _celda(ws, f, 1, "PLAN DE ACCIÓN", negrita=True, tam=9, color=GRIS)
        f += 1
        _cabecera_tabla(ws, f, ["Cuándo", "Acción", "", "", "", ""], [34, 30, 20, 20, 20, 20])
        ws.freeze_panes = None
        f += 1
        ini = f
        for p in plan:
            _celda(ws, f, 1, p.get("cuando", ""), negrita=True, tam=10, color=ORO)
            ws.merge_cells(start_row=f, start_column=2, end_row=f, end_column=6)
            _celda(ws, f, 2, p.get("accion", ""), wrap=True, tam=10)
            ws.row_dimensions[f].height = 34
            f += 1
        _bordear(ws, ini, f - 1, 6)
    return ws


def _hoja_proyeccion(wb, fc, inf):
    """Periodo a periodo, con el rango. Es la hoja que se usa para planificar."""
    ws = wb.create_sheet("Proyección")
    ws.sheet_view.showGridLines = False
    u = inf.get("kpis", {}).get("unidad", "unidades")
    f = _titulo_hoja(ws, "Proyección periodo a periodo",
                     "«Mínimo» y «Máximo» son el rango con 80% de probabilidad: de cada 10 "
                     "periodos, 8 caen dentro. Planifica con el mínimo y ten capacidad para el máximo.",
                     ancho=5)
    _cabecera_tabla(ws, f, ["Serie", "Periodo", f"Proyección ({u})", "Mínimo", "Máximo"],
                    [24, 14, 18, 14, 14])
    f += 1
    ini = f
    d = fc.sort_values(["unique_id", "ds"])
    for r in d.itertuples(index=False):
        _celda(ws, f, 1, str(r.unique_id), tam=10)
        _celda(ws, f, 2, getattr(r, "ds"), tam=10, formato="dd/mm/yyyy")
        _celda(ws, f, 3, float(r.Forecast), negrita=True, tam=10, formato="#,##0")
        _celda(ws, f, 4, float(getattr(r, "Lo_80", 0)), tam=10, color=GRIS, formato="#,##0")
        _celda(ws, f, 5, float(getattr(r, "Hi_80", 0)), tam=10, color=GRIS, formato="#,##0")
        f += 1
    _bordear(ws, ini, f - 1, 5)
    # Total al pie
    _celda(ws, f, 2, "TOTAL", negrita=True, tam=10, color=NAVY)
    _celda(ws, f, 3, float(d["Forecast"].sum()), negrita=True, tam=11,
           color=NAVY, fondo=GRIS_CLARO, formato="#,##0")
    return ws


def _hoja_equipo_vs_modelo(wb, comp):
    ws = wb.create_sheet("Equipo vs modelo")
    ws.sheet_view.showGridLines = False
    f = _titulo_hoja(ws, "Tu equipo vs el modelo",
                     comp.get("mensaje", ""), ancho=5)
    _cabecera_tabla(ws, f, ["Serie", "Periodos", "Error del equipo", "Error del modelo", "Quién acertó más"],
                    [24, 12, 18, 18, 20])
    f += 1
    ini = f
    for s in comp.get("series", []):
        _celda(ws, f, 1, s["serie"], tam=10)
        _celda(ws, f, 2, s["n_periodos"], tam=10, formato="#,##0")
        _celda(ws, f, 3, s["wape_manual"] / 100, tam=10, formato="0.0%")
        _celda(ws, f, 4, s["wape_modelo"] / 100, tam=10, formato="0.0%")
        _celda(ws, f, 5, "Tu equipo" if s["gana"] == "manual" else "El modelo",
               negrita=True, tam=10, color=(TEAL if s["gana"] == "modelo" else ORO))
        f += 1
    _bordear(ws, ini, f - 1, 5)
    f += 1
    ws.merge_cells(start_row=f, start_column=1, end_row=f + 1, end_column=5)
    _celda(ws, f, 1, comp.get("accion", ""), wrap=True, tam=10, fondo=GRIS_CLARO)
    return ws


def _hoja_tecnica(wb, tabla, inf):
    ws = wb.create_sheet("Detalle técnico")
    ws.sheet_view.showGridLines = False
    f = _titulo_hoja(ws, "Detalle técnico",
                     "Comparación de los modelos probados. WAPE: error porcentual medio "
                     "(menor es mejor). Sesgo: positivo sobreestima, negativo subestima.",
                     ancho=4)
    d = tabla.rename(columns={"WAPE_%": "error", "BIAS_%": "sesgo"})
    _cabecera_tabla(ws, f, ["Serie", "Modelo", "Error (WAPE)", "Sesgo"], [24, 22, 14, 14])
    f += 1
    ini = f
    for r in d.itertuples(index=False):
        _celda(ws, f, 1, str(r.serie), tam=10)
        _celda(ws, f, 2, str(r.modelo), tam=10)
        _celda(ws, f, 3, float(r.error) / 100, tam=10, formato="0.0%")
        _celda(ws, f, 4, float(r.sesgo) / 100, tam=10, formato="0.0%")
        f += 1
    _bordear(ws, ini, f - 1, 4)

    prec = inf.get("precision")
    if prec:
        f += 2
        _celda(ws, f, 1, "ACIERTO YA VERIFICADO", negrita=True, tam=9, color=GRIS)
        f += 1
        ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=4)
        _celda(ws, f, 1, prec.get("frase", ""), wrap=True, tam=10, fondo=GRIS_CLARO)
        ws.row_dimensions[f].height = 30
    return ws


def construir_libro(fc, tabla, inf):
    """Devuelve un openpyxl.Workbook listo para guardar."""
    from openpyxl import Workbook
    wb = Workbook()
    _hoja_resumen(wb, inf)
    _hoja_proyeccion(wb, fc, inf)
    comp = inf.get("fva")
    if comp and comp.get("series"):
        _hoja_equipo_vs_modelo(wb, comp)
    _hoja_tecnica(wb, tabla, inf)
    wb.active = 0          # abre siempre en el resumen
    return wb


if __name__ == "__main__":  # ponytail: check de la lógica no trivial
    import io as _io
    import pandas as pd
    from openpyxl import load_workbook
    fc = pd.DataFrame({"unique_id": ["A"] * 3,
                       "ds": pd.date_range("2026-01-01", periods=3, freq="MS"),
                       "Forecast": [100.0, 110.0, 120.0],
                       "Lo_80": [90.0, 99.0, 108.0], "Hi_80": [110.0, 121.0, 132.0]})
    tabla = pd.DataFrame({"serie": ["A", "A"], "modelo": ["AutoETS", "SeasonalNaive"],
                          "WAPE_%": [5.2, 11.8], "BIAS_%": [-1.1, 4.0]})
    inf = {"kpis": {"unidad": "contenedores", "horizonte": 3, "frecuencia_plural": "mensuales",
                    "total_fmt": "330", "piso_fmt": "297", "techo_fmt": "363",
                    "variacion_pct": -14.6, "wape_pct": 5.2, "confianza": "alta", "n_series": 1},
           "asistente": {"veredicto": {"frase": "Tu operación cae 15%."},
                         "decisiones": [{"titulo": "Compromete 297", "prioridad": "alta",
                                         "tono": "alerta", "accion": "Fija metas en 297.",
                                         "importa": "Comprometer de más cuesta caro."}],
                         "plan_accion": [{"cuando": "2026-02-01", "accion": "Abrir turno."}]},
           "fva": {"mensaje": "El modelo mejora.", "accion": "Usa el modelo.",
                   "series": [{"serie": "A", "n_periodos": 12, "wape_manual": 12.1,
                               "wape_modelo": 6.1, "gana": "modelo"}]},
           "precision": {"frase": "Mi último pronóstico falló 8%."}}
    buf = _io.BytesIO()
    construir_libro(fc, tabla, inf).save(buf)
    buf.seek(0)
    wb = load_workbook(buf)
    assert wb.sheetnames == ["Resumen", "Proyección", "Equipo vs modelo", "Detalle técnico"], wb.sheetnames
    r = wb["Resumen"]
    textos = " ".join(str(c.value) for row in r.iter_rows() for c in row if c.value)
    assert "Tu operación cae 15%" in textos          # el veredicto abre el informe
    assert "Compromete 297" in textos and "ALTA" in textos
    assert "Comprometer de más cuesta caro" in textos    # el "por qué importa" viaja
    p = wb["Proyección"]
    assert p.freeze_panes is not None                 # cabecera fija al desplazar
    # El total al pie debe cuadrar con la suma
    vals = [c.value for row in p.iter_rows() for c in row
            if isinstance(c.value, (int, float)) and c.value == 330]
    assert vals, "falta el total 330"
    print("excel OK:", wb.sheetnames)
